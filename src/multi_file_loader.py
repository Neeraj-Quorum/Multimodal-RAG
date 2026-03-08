"""
Multi-format file loader for different document types
Supports: PDF, DOCX, DOC, HTML, Images, Code, TXT, CSV, Excel, SQL, XML, JSON, YAML, Config, and more
"""

from pathlib import Path
import logging
import hashlib
import base64
import os
from typing import List, Dict, Any, Optional
from datetime import datetime

# Document processing libraries
from pypdf import PdfReader
from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from PIL import Image
import io
import csv
import json
import warnings

logger = logging.getLogger(__name__)

# Maximum file size to process (200 MB) — larger files are skipped with a warning
MAX_FILE_SIZE_BYTES = 200 * 1024 * 1024

# Directories to always skip (e.g. .git internals, build outputs)
SKIP_DIRS = {'.git', '__pycache__', 'node_modules', '.svn', '.hg'}


class MultiFileLoader:
    """
    Load documents from multiple file formats:
    - PDF (text + embedded images)
    - DOCX (text + embedded images)
    - DOC (legacy Word binary — best-effort via textract/antiword, fallback to skip)
    - HTML/HTM (text + linked images)
    - PNG/JPG/JPEG/GIF/BMP/ICO (images)
    - TXT/MD/RST/MARKDOWN (plain text)
    - CSV (tabular data)
    - XLSX/XLS/XLSM (Excel spreadsheets)
    - SQL (SQL scripts)
    - XML/XSLT/XSD (XML-based files)
    - JSON (JSON data)
    - YML/YAML (YAML config)
    - Code files (.py, .js, .java, .cpp, .c, .h, .hpp, .cs, .go, .rs, .rb, .php,
      .ts, .tsx, .jsx, .pas, .inc, .sj, .ps1, .psm1, .psd1, .cmd, .bat, .sh)
    - Config files (.ini, .cfg, .config, .env, .properties, .toml)
    - Project files (.sln, .csproj, .cbproj, .groupproj, .dfm, .rc, .resx)
    - Log files (.log)
    - Report definition files (.css)
    """

    SUPPORTED_EXTENSIONS = {
        'pdf': ['.pdf'],
        'docx': ['.docx'],
        'doc_legacy': ['.doc'],
        'html': ['.html', '.htm'],
        'image': ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.ico'],
        'text': ['.txt', '.md', '.rst', '.markdown', '.log'],
        'csv': ['.csv'],
        'excel': ['.xlsx', '.xls', '.xlsm'],
        'code': [
            '.py', '.js', '.java', '.cpp', '.c', '.h', '.hpp', '.cs', '.go',
            '.rs', '.rb', '.php', '.ts', '.tsx', '.jsx',
            # Delphi / Pascal / C++ Builder
            '.pas', '.inc', '.dfm',
            # SmartBear TestComplete JavaScript
            '.sj',
            # PowerShell
            '.ps1', '.psm1', '.psd1',
            # Shell / batch scripts
            '.cmd', '.bat', '.sh',
            # Resource / project descriptors (text-based XML)
            '.rc', '.resx', '.csproj', '.cbproj', '.sln', '.groupproj',
            # Web assets
            '.css',
        ],
        'structured_text': [
            '.sql', '.xml', '.xslt', '.xsd', '.json', '.yml', '.yaml',
            '.ini', '.cfg', '.config', '.env', '.properties', '.toml',
            '.dotx',
        ],
    }

    def __init__(self):
        """Initialize the loader"""
        self.loaded_files = []
        self.skipped_files: List[Dict[str, str]] = []
        self.file_stats = {
            'pdf': 0, 'docx': 0, 'doc_legacy': 0, 'html': 0,
            'image': 0, 'text': 0, 'csv': 0, 'excel': 0,
            'code': 0, 'structured_text': 0, 'total': 0
        }
        self.error_files: List[Dict[str, str]] = []
        logger.info("✓ MultiFileLoader initialized")
    
    def _generate_file_id(self, filepath: Path) -> str:
        """Generate unique file ID based on file path hash"""
        file_hash = hashlib.sha256(str(filepath).encode()).hexdigest()[:12]
        return f"file_{file_hash}"

    def _get_file_type(self, filepath: Path) -> str:
        """Determine file type from extension"""
        ext = filepath.suffix.lower()
        for file_type, extensions in self.SUPPORTED_EXTENSIONS.items():
            if ext in extensions:
                return file_type
        return 'unknown'

    def _is_safe_to_load(self, filepath: Path) -> Optional[str]:
        """
        Pre-flight checks before loading a file.
        Returns None if safe, or a reason string if the file should be skipped.
        """
        try:
            size = filepath.stat().st_size
        except OSError as e:
            return f"Cannot stat file: {e}"

        if size == 0:
            return "Zero-byte file"

        if size > MAX_FILE_SIZE_BYTES:
            return f"File too large ({size / (1024**2):.0f} MB > {MAX_FILE_SIZE_BYTES / (1024**2):.0f} MB limit)"

        return None

    def _read_text_with_fallback(self, filepath: Path) -> str:
        """Read a text file trying multiple encodings to avoid data loss."""
        for encoding in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252'):
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    return f.read()
            except (UnicodeDecodeError, UnicodeError):
                continue
        # Last resort: read as bytes and decode replacing errors
        with open(filepath, 'rb') as f:
            return f.read().decode('utf-8', errors='replace')
    
    def load_pdf(self, filepath: Path) -> Dict[str, Any]:
        """
        Load text from PDF using pypdf (fast, no Tesseract/OCR needed).
        Returns: dict with text_chunks extracted from each page.
        """
        logger.info(f"Loading PDF: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping PDF {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        try:
            with warnings.catch_warnings():
                warnings.simplefilter('ignore')
                reader = PdfReader(str(filepath))

            file_id = self._generate_file_id(filepath)
            text_chunks = []
            total_text = []

            for page_num, page in enumerate(reader.pages):
                try:
                    text = page.extract_text() or ''
                except Exception:
                    text = ''
                if text.strip():
                    total_text.append(text)
                    text_chunks.append({
                        'element_id': f"{file_id}_page_{page_num}",
                        'content': text,
                        'type': 'text',
                        'page': page_num + 1
                    })

            # If no text was extracted from any page, record it but don't error
            if not text_chunks:
                # Possibly a scanned/image-only PDF — store as placeholder
                text_chunks.append({
                    'element_id': f"{file_id}_page_0",
                    'content': f"[PDF with {len(reader.pages)} page(s) — no extractable text (possibly scanned/image-only)]",
                    'type': 'text',
                    'page': 0
                })

            self.file_stats['pdf'] += 1
            self.file_stats['total'] += 1

            return {
                'file_id': file_id,
                'filename': filepath.name,
                'filepath': str(filepath),
                'type': 'pdf',
                'text_chunks': text_chunks,
                'tables': [],
                'images': [],
                'page_count': len(reader.pages),
                'element_count': len(text_chunks),
                'loaded_at': datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Error loading PDF {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name, 'filepath': str(filepath)}
    
    def load_docx(self, filepath: Path) -> Dict[str, Any]:
        """
        Load text and images from DOCX (Office Open XML only).
        NOTE: .doc (legacy binary) is handled separately by load_doc_legacy().
        """
        logger.info(f"Loading DOCX: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping DOCX {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}
        
        try:
            doc = DocxDocument(filepath)
            file_id = self._generate_file_id(filepath)
            
            text_chunks = []
            images = []
            
            # Extract paragraphs
            for idx, para in enumerate(doc.paragraphs):
                if para.text.strip():
                    text_chunks.append({
                        'element_id': f"{file_id}_text_{idx}",
                        'content': para.text,
                        'type': 'text',
                        'style': para.style.name if para.style else 'Normal'
                    })
            
            # Extract images from document
            for idx, rel in enumerate(doc.part.rels.values()):
                if "image" in rel.target_ref:
                    try:
                        image_data = rel.target_part.blob
                        image_base64 = base64.b64encode(image_data).decode('utf-8')
                        images.append({
                            'element_id': f"{file_id}_image_{idx}",
                            'content': image_base64,
                            'type': 'image',
                            'format': rel.target_ref.split('.')[-1]
                        })
                    except Exception as img_error:
                        logger.warning(f"Could not extract image {idx}: {img_error}")
            
            self.file_stats['docx'] += 1
            self.file_stats['total'] += 1
            
            return {
                'file_id': file_id,
                'filename': filepath.name,
                'filepath': str(filepath),
                'type': 'docx',
                'text_chunks': text_chunks,
                'images': images,
                'element_count': len(text_chunks) + len(images),
                'loaded_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error loading DOCX {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name}

    def load_doc_legacy(self, filepath: Path) -> Dict[str, Any]:
        """
        Load legacy .doc (binary Word) files.
        python-docx does NOT support .doc — we use textract or antiword if available,
        otherwise fall back to raw binary-to-text extraction.
        """
        logger.info(f"Loading legacy DOC: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping DOC {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        file_id = self._generate_file_id(filepath)

        # Try textract first (if installed)
        try:
            import textract
            raw_bytes = textract.process(str(filepath))
            content = raw_bytes.decode('utf-8', errors='replace')
        except ImportError:
            # textract not available — do a best-effort binary string extraction
            logger.warning(f"textract not installed; extracting readable strings from .doc: {filepath.name}")
            try:
                with open(filepath, 'rb') as f:
                    raw = f.read()
                # Extract printable ASCII runs longer than 4 chars
                import re
                content = '\n'.join(re.findall(rb'[\x20-\x7E]{4,}', raw, re.DOTALL)
                                    .__iter__().__class__(
                                        s.decode('ascii', errors='replace') for s in re.findall(rb'[\x20-\x7E]{4,}', raw)))
            except Exception:
                # Absolute fallback
                with open(filepath, 'rb') as f:
                    raw = f.read()
                import re
                strings = [m.group().decode('ascii', errors='replace')
                           for m in re.finditer(rb'[\x20-\x7E]{4,}', raw)]
                content = '\n'.join(strings)
        except Exception as e:
            logger.error(f"Error extracting .doc {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name, 'filepath': str(filepath)}

        if not content.strip():
            return {'error': 'No readable text extracted from .doc', 'filename': filepath.name, 'filepath': str(filepath)}

        chunks = self._chunk_text(file_id, content)

        self.file_stats['doc_legacy'] += 1
        self.file_stats['total'] += 1

        return {
            'file_id': file_id,
            'filename': filepath.name,
            'filepath': str(filepath),
            'type': 'doc_legacy',
            'text_chunks': chunks,
            'element_count': len(chunks),
            'loaded_at': datetime.now().isoformat()
        }

    def load_csv(self, filepath: Path) -> Dict[str, Any]:
        """
        Load CSV files as tabular data.
        Each row becomes a text chunk, with headers preserved as context.
        """
        logger.info(f"Loading CSV: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping CSV {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        try:
            content = self._read_text_with_fallback(filepath)
            file_id = self._generate_file_id(filepath)

            # Parse CSV
            reader = csv.reader(io.StringIO(content))
            rows = list(reader)

            if not rows:
                return {'error': 'Empty CSV', 'filename': filepath.name, 'filepath': str(filepath)}

            headers = rows[0] if rows else []
            text_chunks = []

            # Store header row
            text_chunks.append({
                'element_id': f"{file_id}_header",
                'content': f"CSV Headers: {', '.join(headers)}",
                'type': 'table_header',
            })

            # Chunk rows in batches of 50 to keep context manageable
            batch_size = 50
            data_rows = rows[1:]
            for batch_idx in range(0, len(data_rows), batch_size):
                batch = data_rows[batch_idx:batch_idx + batch_size]
                batch_text_lines = []
                for row in batch:
                    row_dict = {h: v for h, v in zip(headers, row)} if headers else {}
                    batch_text_lines.append(
                        ', '.join(f"{k}: {v}" for k, v in row_dict.items()) if row_dict
                        else ', '.join(row)
                    )
                text_chunks.append({
                    'element_id': f"{file_id}_rows_{batch_idx}",
                    'content': '\n'.join(batch_text_lines),
                    'type': 'table',
                    'row_range': f"{batch_idx + 1}-{batch_idx + len(batch)}"
                })

            self.file_stats['csv'] += 1
            self.file_stats['total'] += 1

            return {
                'file_id': file_id,
                'filename': filepath.name,
                'filepath': str(filepath),
                'type': 'csv',
                'text_chunks': text_chunks,
                'row_count': len(data_rows),
                'column_count': len(headers),
                'element_count': len(text_chunks),
                'loaded_at': datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Error loading CSV {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name, 'filepath': str(filepath)}

    def load_excel(self, filepath: Path) -> Dict[str, Any]:
        """
        Load Excel files (.xlsx, .xls, .xlsm) using openpyxl.
        Falls back to pandas if openpyxl fails on .xls.
        """
        logger.info(f"Loading Excel: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping Excel {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        file_id = self._generate_file_id(filepath)
        text_chunks = []

        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                headers = [str(c) if c is not None else '' for c in rows[0]]
                data_rows = rows[1:]

                # Chunk by batches of 50 rows
                batch_size = 50
                for batch_idx in range(0, max(len(data_rows), 1), batch_size):
                    batch = data_rows[batch_idx:batch_idx + batch_size]
                    lines = []
                    for row in batch:
                        row_dict = {h: str(v) if v is not None else '' for h, v in zip(headers, row)}
                        lines.append(', '.join(f"{k}: {v}" for k, v in row_dict.items() if v))
                    text_chunks.append({
                        'element_id': f"{file_id}_{sheet_name}_{batch_idx}",
                        'content': f"Sheet: {sheet_name}\nHeaders: {', '.join(headers)}\n" + '\n'.join(lines),
                        'type': 'table',
                        'sheet': sheet_name,
                    })
            wb.close()

        except ImportError:
            logger.warning("openpyxl not installed; trying pandas for Excel")
            try:
                import pandas as pd
                xls = pd.ExcelFile(str(filepath))
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=sheet_name)
                    text_chunks.append({
                        'element_id': f"{file_id}_{sheet_name}_0",
                        'content': f"Sheet: {sheet_name}\n{df.to_string(max_rows=200)}",
                        'type': 'table',
                        'sheet': sheet_name,
                    })
            except Exception as e2:
                logger.error(f"Error loading Excel with pandas {filepath}: {e2}")
                return {'error': str(e2), 'filename': filepath.name, 'filepath': str(filepath)}

        except Exception as e:
            # For old .xls format, openpyxl will fail; try pandas
            if filepath.suffix.lower() == '.xls':
                try:
                    import pandas as pd
                    xls = pd.ExcelFile(str(filepath))
                    for sheet_name in xls.sheet_names:
                        df = pd.read_excel(xls, sheet_name=sheet_name)
                        text_chunks.append({
                            'element_id': f"{file_id}_{sheet_name}_0",
                            'content': f"Sheet: {sheet_name}\n{df.to_string(max_rows=200)}",
                            'type': 'table',
                            'sheet': sheet_name,
                        })
                except Exception as e2:
                    logger.error(f"Error loading .xls {filepath}: {e2}")
                    return {'error': str(e2), 'filename': filepath.name, 'filepath': str(filepath)}
            else:
                logger.error(f"Error loading Excel {filepath}: {e}")
                return {'error': str(e), 'filename': filepath.name, 'filepath': str(filepath)}

        if not text_chunks:
            return {'error': 'No data found in Excel file', 'filename': filepath.name, 'filepath': str(filepath)}

        self.file_stats['excel'] += 1
        self.file_stats['total'] += 1

        return {
            'file_id': file_id,
            'filename': filepath.name,
            'filepath': str(filepath),
            'type': 'excel',
            'text_chunks': text_chunks,
            'element_count': len(text_chunks),
            'loaded_at': datetime.now().isoformat()
        }

    def load_structured_text(self, filepath: Path) -> Dict[str, Any]:
        """
        Load structured text files: SQL, XML, JSON, YAML, INI, config, etc.
        These are text-readable and loaded similarly to plain text but tagged
        with their specific sub-type for downstream processing.
        """
        logger.info(f"Loading structured text: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        try:
            content = self._read_text_with_fallback(filepath)
            file_id = self._generate_file_id(filepath)

            ext = filepath.suffix.lower()
            sub_type = ext.lstrip('.')

            chunks = self._chunk_text(file_id, content)

            self.file_stats['structured_text'] += 1
            self.file_stats['total'] += 1

            return {
                'file_id': file_id,
                'filename': filepath.name,
                'filepath': str(filepath),
                'type': 'structured_text',
                'sub_type': sub_type,
                'text_chunks': chunks,
                'element_count': len(chunks),
                'loaded_at': datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Error loading structured text {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name, 'filepath': str(filepath)}

    def _chunk_text(self, file_id: str, content: str, max_chunk_size: int = 2000) -> List[Dict[str, Any]]:
        """Split text content into manageable chunks."""
        chunks = []
        if len(content) > max_chunk_size:
            paragraphs = content.split('\n\n')
            current_chunk = []
            current_len = 0
            chunk_idx = 0
            for para in paragraphs:
                if para.strip():
                    if current_len + len(para) > max_chunk_size and current_chunk:
                        chunks.append({
                            'element_id': f"{file_id}_text_{chunk_idx}",
                            'content': '\n\n'.join(current_chunk),
                            'type': 'text'
                        })
                        chunk_idx += 1
                        current_chunk = []
                        current_len = 0
                    current_chunk.append(para)
                    current_len += len(para)
            if current_chunk:
                chunks.append({
                    'element_id': f"{file_id}_text_{chunk_idx}",
                    'content': '\n\n'.join(current_chunk),
                    'type': 'text'
                })
        else:
            chunks.append({
                'element_id': f"{file_id}_text_0",
                'content': content,
                'type': 'text'
            })
        return chunks

    def load_html(self, filepath: Path) -> Dict[str, Any]:
        """
        Load HTML using BeautifulSoup (fast, no external dependencies).
        Returns: dict with text_chunks extracted from the HTML body.
        """
        logger.info(f"Loading HTML: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping HTML {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        try:
            raw = self._read_text_with_fallback(filepath)
            soup = BeautifulSoup(raw, 'html.parser')

            # Remove script and style elements
            for tag in soup(['script', 'style', 'noscript']):
                tag.decompose()

            file_id = self._generate_file_id(filepath)
            text = soup.get_text(separator='\n', strip=True)

            chunks = self._chunk_text(file_id, text) if text.strip() else []

            if not chunks:
                chunks.append({
                    'element_id': f"{file_id}_text_0",
                    'content': '[HTML file with no extractable text]',
                    'type': 'text'
                })

            self.file_stats['html'] += 1
            self.file_stats['total'] += 1

            return {
                'file_id': file_id,
                'filename': filepath.name,
                'filepath': str(filepath),
                'type': 'html',
                'text_chunks': chunks,
                'element_count': len(chunks),
                'loaded_at': datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Error loading HTML {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name, 'filepath': str(filepath)}
    
    def load_image(self, filepath: Path) -> Dict[str, Any]:
        """
        Load standalone image
        Returns: dict with image data
        """
        logger.info(f"Loading Image: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping image {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        try:
            with Image.open(filepath) as img:
                # Convert to RGB if needed
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                
                # Convert to base64
                buffered = io.BytesIO()
                img.save(buffered, format="PNG")
                image_base64 = base64.b64encode(buffered.getvalue()).decode('utf-8')
                
                file_id = self._generate_file_id(filepath)
                
                self.file_stats['image'] += 1
                self.file_stats['total'] += 1
                
                return {
                    'file_id': file_id,
                    'filename': filepath.name,
                    'filepath': str(filepath),
                    'type': 'image',
                    'image_data': image_base64,
                    'dimensions': f"{img.width}x{img.height}",
                    'format': img.format,
                    'element_count': 1,
                    'loaded_at': datetime.now().isoformat()
                }
                
        except Exception as e:
            logger.error(f"Error loading image {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name}
    
    def load_code(self, filepath: Path) -> Dict[str, Any]:
        """
        Load code files
        Returns: dict with code content
        """
        logger.info(f"Loading Code: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping code {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        try:
            content = self._read_text_with_fallback(filepath)
            
            file_id = self._generate_file_id(filepath)
            
            self.file_stats['code'] += 1
            self.file_stats['total'] += 1
            
            return {
                'file_id': file_id,
                'filename': filepath.name,
                'filepath': str(filepath),
                'type': 'code',
                'language': filepath.suffix[1:],  # Remove the dot
                'content': content,
                'lines': len(content.split('\n')),
                'element_count': 1,
                'loaded_at': datetime.now().isoformat()
            }
            
        except Exception as e:
            logger.error(f"Error loading code {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name}
    
    def load_text(self, filepath: Path) -> Dict[str, Any]:
        """
        Load plain text files
        Returns: dict with text content
        """
        logger.info(f"Loading Text: {filepath.name}")
        skip_reason = self._is_safe_to_load(filepath)
        if skip_reason:
            logger.warning(f"Skipping text {filepath.name}: {skip_reason}")
            return {'error': skip_reason, 'filename': filepath.name, 'filepath': str(filepath)}

        try:
            content = self._read_text_with_fallback(filepath)
            file_id = self._generate_file_id(filepath)

            chunks = self._chunk_text(file_id, content)

            self.file_stats['text'] += 1
            self.file_stats['total'] += 1

            return {
                'file_id': file_id,
                'filename': filepath.name,
                'filepath': str(filepath),
                'type': 'text',
                'text_chunks': chunks,
                'element_count': len(chunks),
                'loaded_at': datetime.now().isoformat()
            }

        except Exception as e:
            logger.error(f"Error loading text {filepath}: {e}")
            return {'error': str(e), 'filename': filepath.name, 'filepath': str(filepath)}
    
    def load_file(self, filepath: Path) -> Dict[str, Any]:
        """
        Load any supported file type automatically.
        Dispatches to the correct loader based on file extension.
        """
        file_type = self._get_file_type(filepath)

        loaders = {
            'pdf': self.load_pdf,
            'docx': self.load_docx,
            'doc_legacy': self.load_doc_legacy,
            'html': self.load_html,
            'image': self.load_image,
            'code': self.load_code,
            'text': self.load_text,
            'csv': self.load_csv,
            'excel': self.load_excel,
            'structured_text': self.load_structured_text,
        }

        loader = loaders.get(file_type)
        if loader:
            return loader(filepath)
        else:
            logger.warning(f"Unsupported file type: {filepath.suffix} -> {filepath}")
            self.skipped_files.append({'filename': filepath.name, 'filepath': str(filepath), 'reason': f'Unsupported extension: {filepath.suffix}'})
            return {'error': f'Unsupported file type: {filepath.suffix}', 'filename': filepath.name, 'filepath': str(filepath)}
    
    def load_all_documents(self, folder_path: Path, recursive: bool = True) -> List[Dict[str, Any]]:
        """
        Load all documents from a folder.
        Args:
            folder_path: Path to folder
            recursive: Whether to scan subfolders
        Returns:
            List of document dictionaries
        """
        # Suppress noisy library warnings (openpyxl dates, PIL palette, etc.)
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        warnings.filterwarnings('ignore', category=UserWarning, module='PIL')

        documents = []
        folder_path = Path(folder_path)

        logger.info(f"\n{'='*60}")
        logger.info(f"Scanning folder: {folder_path}")
        logger.info(f"{'='*60}")

        if not folder_path.exists():
            logger.error(f"Folder does not exist: {folder_path}")
            return documents

        # Collect files, skipping excluded directories
        all_files = []
        for root, dirs, files in os.walk(folder_path):
            # Remove excluded dirs so os.walk won't descend into them
            dirs[:] = [d for d in dirs if d not in SKIP_DIRS]
            if not recursive and Path(root) != folder_path:
                continue
            for f in files:
                all_files.append(Path(root) / f)

        # Separate supported vs unsupported
        supported_files = []
        unsupported_files = []
        for f in all_files:
            if self._get_file_type(f) != 'unknown':
                supported_files.append(f)
            else:
                unsupported_files.append(f)
                self.skipped_files.append({
                    'filename': f.name,
                    'filepath': str(f),
                    'reason': f'Unsupported extension: {f.suffix}'
                })

        logger.info(f"\nFound {len(all_files)} total files")
        logger.info(f"  Supported: {len(supported_files)}")
        logger.info(f"  Unsupported (will skip): {len(unsupported_files)}")

        # Load each supported file
        loaded = 0
        errors = 0
        for idx, filepath in enumerate(supported_files):
            if (idx + 1) % 500 == 0:
                logger.info(f"  Progress: {idx + 1}/{len(supported_files)} files processed...")
            doc = self.load_file(filepath)
            if 'error' not in doc:
                documents.append(doc)
                self.loaded_files.append(filepath)
                loaded += 1
            else:
                errors += 1
                self.error_files.append({
                    'filename': filepath.name,
                    'filepath': str(filepath),
                    'error': doc.get('error', 'Unknown error')
                })

        logger.info(f"\nLoaded: {loaded}, Errors: {errors}, Skipped: {len(unsupported_files)}")

        # Print statistics
        self._print_stats()

        return documents
    
    def _print_stats(self):
        """Print loading statistics"""
        logger.info(f"\n{'='*60}")
        logger.info("LOADING STATISTICS")
        logger.info(f"{'='*60}")
        for ftype, count in sorted(self.file_stats.items()):
            if ftype != 'total' and count > 0:
                logger.info(f"{ftype:20s}: {count}")
        logger.info(f"{'='*60}")
        logger.info(f"{'TOTAL':20s}: {self.file_stats['total']} files")
        if self.skipped_files:
            logger.info(f"{'SKIPPED':20s}: {len(self.skipped_files)} files")
        if self.error_files:
            logger.info(f"{'ERRORS':20s}: {len(self.error_files)} files")
        logger.info(f"{'='*60}\n")

    def get_loading_report(self) -> Dict[str, Any]:
        """Return a summary report of loading results for inspection."""
        return {
            'stats': dict(self.file_stats),
            'loaded_count': len(self.loaded_files),
            'skipped_count': len(self.skipped_files),
            'error_count': len(self.error_files),
            'skipped_files': self.skipped_files[:50],  # First 50 for brevity
            'error_files': self.error_files[:50],
        }
